"""
Excel export module for comparison results
Экспорт результатов сравнения в Excel - Универсальная версия
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
from typing import List, Dict
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ResultsExporter:
    """Export generic comparison results to Excel"""
    
    COLORS = {
        'high': 'C6EFCE',      # Green (>80%)
        'medium': 'FFEB9C',    # Yellow (50-80%)
        'low': 'FFC7CE',       # Red (<50%)
        'header': '333333',    # Dark grey
        'subheader': '666666', # Medium grey
    }
    
    def __init__(self):
        self.wb = None
    
    def export_results(self, results: List[Dict], new_filename: str, output_path: str):
        self.wb = Workbook()
        self._create_summary_sheet(results, new_filename)
        self.wb.save(output_path)
        logger.info(f"Exported results to {output_path}")
    
    def _create_summary_sheet(self, results: List[Dict], new_filename: str):
        ws = self.wb.active
        ws.title = "Результаты сравнения"
        ws.sheet_view.showGridLines = False
        
        # Title
        ws.merge_cells('B2:H2')
        ws['B2'] = f"Результаты сравнения: {new_filename}"
        ws['B2'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['B2'].fill = PatternFill(start_color=self.COLORS['header'], 
                                     end_color=self.COLORS['header'], fill_type="solid")
        ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 30
        
        # Date
        ws['B3'] = f"Дата сравнения: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['B3'].font = Font(size=10, italic=True, color="666666")
        
        # Headers
        headers = [
            'Место', 'Имя файла архива', 'Дата добавления', 'Общий % схожести',
            '% Совпадения содержимого строк', '% Совпадения структуры колонок', 
            'Кол-во позиций'
        ]
        
        header_row = 5
        for col, header in enumerate(headers, 2):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLORS['subheader'],
                                   end_color=self.COLORS['subheader'], fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws.row_dimensions[header_row].height = 40
        
        # Data rows
        for idx, result in enumerate(results, 1):
            row = header_row + idx
            sim = result['similarity']
            
            total_sim = sim['total']
            if total_sim >= 80: fill_color = self.COLORS['high']
            elif total_sim >= 50: fill_color = self.COLORS['medium']
            else: fill_color = self.COLORS['low']
            
            data = [
                idx,  
                result['filename'],  
                result.get('upload_date', ''),  
                f"{sim['total']}%",  
                f"{sim.get('content', 0)}%",  
                f"{sim.get('columns', 0)}%",  
                result.get('total_items', 0)
            ]
            
            for col, value in enumerate(data, 2):
                cell = ws.cell(row=row, column=col, value=value)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if col in [3, 5]:  
                    cell.font = Font(bold=True)
        
        # Adjust column widths
        column_widths = {
            'B': 8,   
            'C': 45,  
            'D': 18,  
            'E': 18,  
            'F': 30,  
            'G': 30,  
            'H': 15,  
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
